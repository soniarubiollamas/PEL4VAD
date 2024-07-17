import numpy as np
import os
from openpyxl import load_workbook

def create_original_gt(gt_file, duration_dict):
    with open(gt_file, 'r') as f:
        for line in f:
            # Extract video name and ground truth information
            video_name, anomaly, start_frame, end_frame, second_start_frame, second_end_frame = line.strip().split()
            start_frame, end_frame, second_start_frame, second_end_frame = int(start_frame), int(end_frame), int(second_start_frame), int(second_end_frame)

            video_name = video_name.split(".")[0]
            video_duration = duration_dict.get(video_name)
            if video_duration is None:
                print(f"Warning: Duration not found for video: {video_name}")
                continue
            total_frames = int(video_duration*30) # 30 cause there are 30 fps
            video_predictions = np.zeros(total_frames)

            if start_frame != -1:
                if start_frame == end_frame:
                   video_predictions[start_frame-1] = 1
                else:
                    video_predictions[start_frame-1:end_frame-1] = 1
            if second_start_frame != -1:
                if second_start_frame == second_end_frame:
                   video_predictions[second_start_frame-1] = 1
                else: 
                    video_predictions[second_start_frame-1:second_end_frame] = 1

            np.save('frame_label/gt/'+video_name+'_pred.npy', video_predictions)


def create_duration_dict(excel_file):
  wb = load_workbook(excel_file)  # Load the Excel workbook
  sheet = wb.active  # Assuming data is in the active sheet

  duration_dict = {}
  for row in sheet.iter_rows(min_row=2):  # Skip header row (assuming row 1)
    filename = row[0].value  # Assuming filename is in column 1
    filename = filename.split(".")[0]
    duration = row[5].value  # Assuming duration is in column 6 (index 5)
    if filename and duration:  # Check for valid values
      duration_dict[filename] = duration

  return duration_dict


excel_file = "annotations/buenos resultados/time_prediction_19-05-24.xlsx"  
duration_dict = create_duration_dict(excel_file)

gt_file = "Temporal_Anomaly_Annotation.txt"
create_original_gt(gt_file,duration_dict)
