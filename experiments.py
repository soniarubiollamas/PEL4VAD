import openpyxl
import json

# Function to process a single Excel row
def process_excel_row(row, json_data):
    filename = row[0].value  # Extract filename from the first column
    filename = filename.split('.')[0] # remove the file extension
    if filename in json_data:
        duration = json_data[filename]["duration"]
        row[5].value = duration

# Load the Excel file
excel_file = "annotations/time_prediction.xlsx"
workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook.active

# Load the JSON data
json_file = "annotations/UCFCrime_Train.json"
with open(json_file, 'r') as f:
    json_data = json.load(f)

# Add a header for the duration column (if it doesn't exist)
if worksheet.cell(row=1, column=6).value != "Duration":
    worksheet.cell(row=1, column=6).value = "Duration"

# Iterate through Excel rows, starting from the second row (assuming headers)
for row in worksheet.iter_rows(min_row=2):
    process_excel_row(row, json_data)

# Save the modified Excel file
workbook.save(excel_file)
