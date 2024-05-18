import openpyxl
import json
import os

# Load the Excel file
# create excel if it doesnt exist
excel_file = "annotations/histogram_times.xlsx"
if not os.path.exists(excel_file):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.cell(row=1, column=1).value = "Filename"
    worksheet.cell(row=1, column=2).value = "Duration"
    workbook.save(excel_file)
workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook.active

# Load the JSON data
json_files = ["annotations/UCFCrime_Val.json", "annotations/UCFCrime_Test.json", "annotations/UCFCrime_Train.json"]
# for i in range(3):

#     json_file = json_files[i]
#     with open(json_file, 'r') as f:
#         json_data = json.load(f)

    # # Add a header for the duration column (if it doesn't exist)
    # if worksheet.cell(row=1, column=6).value != "Duration":
    #     worksheet.cell(row=1, column=6).value = "Duration"

    # for filename, data in json_data.items():
    #     # breakpoint()
    #     duration = data["duration"]
    #     # read all the rows in the excel file
    #     # if filename is not in the excel file, append it
    #     breakpoint()

    #     for row in worksheet.iter_rows(min_row=2):
    #         if row[0].value != filename:
    #             worksheet.append([filename, duration])

    # Load and process JSON data (collect filenames and durations)
filename_duration_map = {}
for json_file in json_files:
    with open(json_file, 'r') as f:
        json_data = json.load(f)
    for filename, data in json_data.items():
        # no "" in "Abuse043_x264"
        # filename = filename.replace('"', '')
        duration = data["duration"]
        filename_duration_map[filename] = duration

    # add the filename and duration to the excel file
    for filename, duration in filename_duration_map.items():
        # if filename is not in the excel file, append it
        if filename not in worksheet:
            worksheet.append([filename, duration])



# Save the modified Excel file
workbook.save(excel_file)
